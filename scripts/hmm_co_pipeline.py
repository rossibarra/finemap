#!/usr/bin/env python3

from __future__ import annotations

import csv
import math
import os
from collections import Counter, defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


INPUT_XLSX = Path("/Users/jeffreyross-ibarra/Downloads/gb-2013-14-9-r103-S4.xlsx")
MATRIX_DIR = Path("results/hmm_cleaned_matrices")
HMM_EVENTS_OUTPUT = Path("results/hmm_co_events_long.tsv")
QC_SUMMARY_OUTPUT = Path("results/hmm_qc_summary.tsv")

MISSINGNESS_MAX = 0.20
ALLELE_BALANCE_MIN = 0.10
ALLELE_BALANCE_MAX = 0.90
ISOLATED_FLIP_MAX = 0.12
SAMPLE_MISSINGNESS_MAX = 0.30
EMISSION_MATCH = 0.98
EMISSION_MISMATCH = 0.02
MISSING_EMISSION = 0.50
TRANSITION_RATE_PER_BP = 1e-8
TRANSITION_MIN = 1e-6
TRANSITION_MAX = 0.05
DOUBLE_CO_MAX_BP = 2_000_000


@dataclass
class Marker:
    chromosome: int
    coordinate: int
    snp_name: str
    genotypes: str


@dataclass
class HMMEvent:
    sample_id: str
    map_name: str
    chromosome: int
    event_index: int
    left_snp: str
    right_snp: str
    left_coordinate: int
    right_coordinate: int
    left_state: str
    right_state: str
    event_bp: int


def load_markers() -> dict[str, list[Marker]]:
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb["Table_S3"]
    header = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    idx = {name: i for i, name in enumerate(header)}

    markers_by_map: dict[str, list[Marker]] = defaultdict(list)

    for row in ws.iter_rows(min_row=4, values_only=True):
        map_name = row[idx["map"]]
        snp_name = row[idx["SNP_name"]]
        chromosome = row[idx["chr_phy"]]
        coordinate = row[idx["coordinate"]]
        raw_data = row[idx["raw_data"]]

        if not map_name or not snp_name or raw_data is None:
            continue
        if chromosome in (None, 0) or coordinate is None:
            continue
        if len(raw_data) <= 6:
            continue

        markers_by_map[map_name].append(
            Marker(
                chromosome=int(chromosome),
                coordinate=int(coordinate),
                snp_name=str(snp_name),
                genotypes=str(raw_data),
            )
        )

    return markers_by_map


def informative_counts(genotypes: str) -> tuple[int, int, int]:
    a_count = genotypes.count("A")
    b_count = genotypes.count("B")
    informative = a_count + b_count
    return a_count, b_count, informative


def choose_best_marker(markers: list[Marker]) -> Marker:
    return max(
        markers,
        key=lambda marker: (
            informative_counts(marker.genotypes)[2],
            -sum(char not in {"A", "B"} for char in marker.genotypes),
            marker.snp_name,
        ),
    )


def deduplicate_markers(markers: list[Marker]) -> list[Marker]:
    grouped: dict[tuple[int, int], list[Marker]] = defaultdict(list)
    for marker in markers:
        grouped[(marker.chromosome, marker.coordinate)].append(marker)
    deduped = [choose_best_marker(group) for group in grouped.values()]
    deduped.sort(key=lambda marker: (marker.chromosome, marker.coordinate, marker.snp_name))
    return deduped


def marker_missingness(genotypes: str, kept_samples: list[int] | None = None) -> float:
    total = 0
    missing = 0
    indices = kept_samples if kept_samples is not None else range(len(genotypes))
    for idx in indices:
        total += 1
        if genotypes[idx] not in {"A", "B"}:
            missing += 1
    return missing / total if total else 1.0


def allele_balance(genotypes: str, kept_samples: list[int] | None = None) -> float | None:
    a_count = 0
    b_count = 0
    indices = kept_samples if kept_samples is not None else range(len(genotypes))
    for idx in indices:
        if genotypes[idx] == "A":
            a_count += 1
        elif genotypes[idx] == "B":
            b_count += 1
    informative = a_count + b_count
    if informative == 0:
        return None
    return a_count / informative


def isolated_flip_rate(markers: list[Marker], marker_idx: int) -> float:
    if marker_idx == 0 or marker_idx == len(markers) - 1:
        return 0.0

    left = markers[marker_idx - 1].genotypes
    middle = markers[marker_idx].genotypes
    right = markers[marker_idx + 1].genotypes

    supported = 0
    isolated = 0
    for sample_idx in range(min(len(left), len(middle), len(right))):
        left_gt = left[sample_idx]
        middle_gt = middle[sample_idx]
        right_gt = right[sample_idx]
        if left_gt not in {"A", "B"} or middle_gt not in {"A", "B"} or right_gt not in {"A", "B"}:
            continue
        if left_gt != right_gt:
            continue
        supported += 1
        if middle_gt != left_gt:
            isolated += 1

    return isolated / supported if supported else 0.0


def compute_sample_missingness(markers: list[Marker]) -> list[float]:
    sample_count = max(len(marker.genotypes) for marker in markers)
    totals = [0] * sample_count
    missings = [0] * sample_count

    for marker in markers:
        for idx in range(sample_count):
            if idx >= len(marker.genotypes):
                totals[idx] += 1
                missings[idx] += 1
                continue
            totals[idx] += 1
            if marker.genotypes[idx] not in {"A", "B"}:
                missings[idx] += 1

    return [missings[idx] / totals[idx] if totals[idx] else 1.0 for idx in range(sample_count)]


def subset_marker(marker: Marker, kept_samples: list[int]) -> Marker:
    return Marker(
        chromosome=marker.chromosome,
        coordinate=marker.coordinate,
        snp_name=marker.snp_name,
        genotypes="".join(marker.genotypes[idx] if idx < len(marker.genotypes) else "-" for idx in kept_samples),
    )


def clean_population(markers: list[Marker]) -> tuple[list[Marker], list[int], Counter]:
    summary = Counter()
    deduped = deduplicate_markers(markers)
    summary["input_markers"] = len(markers)
    summary["deduplicated_markers"] = len(deduped)
    summary["removed_duplicate_markers"] = len(markers) - len(deduped)

    sample_missingness = compute_sample_missingness(deduped)
    kept_samples = [idx for idx, miss in enumerate(sample_missingness) if miss <= SAMPLE_MISSINGNESS_MAX]
    summary["input_samples"] = len(sample_missingness)
    summary["kept_samples"] = len(kept_samples)
    summary["removed_samples_missingness"] = len(sample_missingness) - len(kept_samples)

    missingness_filtered = [
        marker for marker in deduped if marker_missingness(marker.genotypes, kept_samples) <= MISSINGNESS_MAX
    ]
    summary["removed_markers_missingness"] = len(deduped) - len(missingness_filtered)

    balance_filtered = []
    for marker in missingness_filtered:
        balance = allele_balance(marker.genotypes, kept_samples)
        if balance is None:
            continue
        if ALLELE_BALANCE_MIN <= balance <= ALLELE_BALANCE_MAX:
            balance_filtered.append(marker)
    summary["removed_markers_allele_balance"] = len(missingness_filtered) - len(balance_filtered)

    balance_filtered.sort(key=lambda marker: (marker.chromosome, marker.coordinate, marker.snp_name))
    flip_filtered = []
    for idx, marker in enumerate(balance_filtered):
        if isolated_flip_rate(balance_filtered, idx) > ISOLATED_FLIP_MAX:
            continue
        flip_filtered.append(marker)
    summary["removed_markers_isolated_flip"] = len(balance_filtered) - len(flip_filtered)

    final_markers = [subset_marker(marker, kept_samples) for marker in flip_filtered]
    summary["final_markers"] = len(final_markers)
    return final_markers, kept_samples, summary


def write_clean_matrix(map_name: str, markers: list[Marker]) -> None:
    MATRIX_DIR.mkdir(parents=True, exist_ok=True)
    if not markers:
        return
    sample_count = len(markers[0].genotypes)
    path = MATRIX_DIR / f"{map_name}.tsv"
    with path.open("w", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(["chromosome", "coordinate", "snp_name", *[f"{map_name}_ind{i + 1:03d}" for i in range(sample_count)]])
        for marker in markers:
            writer.writerow([marker.chromosome, marker.coordinate, marker.snp_name, *list(marker.genotypes)])


def transition_probability(distance_bp: int) -> float:
    value = TRANSITION_RATE_PER_BP * max(distance_bp, 1)
    return min(max(value, TRANSITION_MIN), TRANSITION_MAX)


def emission_log_prob(state: int, observation: str) -> float:
    if observation not in {"A", "B"}:
        return math.log(MISSING_EMISSION)
    if (state == 0 and observation == "A") or (state == 1 and observation == "B"):
        return math.log(EMISSION_MATCH)
    return math.log(EMISSION_MISMATCH)


def viterbi_decode(markers: list[Marker], sample_idx: int) -> list[int]:
    if not markers:
        return []

    scores = [[float("-inf")] * len(markers) for _ in range(2)]
    back = [[0] * len(markers) for _ in range(2)]

    first_obs = markers[0].genotypes[sample_idx]
    for state in (0, 1):
        scores[state][0] = math.log(0.5) + emission_log_prob(state, first_obs)

    for idx in range(1, len(markers)):
        obs = markers[idx].genotypes[sample_idx]
        distance_bp = markers[idx].coordinate - markers[idx - 1].coordinate
        switch_prob = transition_probability(distance_bp)
        stay_prob = 1.0 - switch_prob
        transition_logs = {
            (0, 0): math.log(stay_prob),
            (0, 1): math.log(switch_prob),
            (1, 0): math.log(switch_prob),
            (1, 1): math.log(stay_prob),
        }

        for state in (0, 1):
            candidates = [
                (scores[prev_state][idx - 1] + transition_logs[(prev_state, state)], prev_state)
                for prev_state in (0, 1)
            ]
            best_score, best_prev = max(candidates, key=lambda item: item[0])
            scores[state][idx] = best_score + emission_log_prob(state, obs)
            back[state][idx] = best_prev

    final_state = 0 if scores[0][-1] >= scores[1][-1] else 1
    path = [final_state]
    for idx in range(len(markers) - 1, 0, -1):
        final_state = back[final_state][idx]
        path.append(final_state)
    path.reverse()
    return path


def collapse_close_double_cos(events: list[HMMEvent]) -> list[HMMEvent]:
    keep = [True] * len(events)
    idx = 0
    while idx < len(events) - 1:
        current = events[idx]
        nxt = events[idx + 1]
        if nxt.event_bp - current.event_bp <= DOUBLE_CO_MAX_BP:
            keep[idx] = False
            keep[idx + 1] = False
            idx += 2
            continue
        idx += 1
    return [event for event, keep_flag in zip(events, keep) if keep_flag]


def build_hmm_events(map_name: str, markers: list[Marker]) -> list[HMMEvent]:
    events: list[HMMEvent] = []
    markers_by_chr: dict[int, list[Marker]] = defaultdict(list)
    for marker in markers:
        markers_by_chr[marker.chromosome].append(marker)

    sample_count = len(markers[0].genotypes) if markers else 0
    for sample_idx in range(sample_count):
        sample_id = f"{map_name}_ind{sample_idx + 1:03d}"
        sample_events: list[HMMEvent] = []

        for chromosome, chromosome_markers in sorted(markers_by_chr.items()):
            decoded = viterbi_decode(chromosome_markers, sample_idx)
            if not decoded:
                continue

            for idx in range(1, len(decoded)):
                if decoded[idx] == decoded[idx - 1]:
                    continue
                left_marker = chromosome_markers[idx - 1]
                right_marker = chromosome_markers[idx]
                left_state = "A" if decoded[idx - 1] == 0 else "B"
                right_state = "A" if decoded[idx] == 0 else "B"
                sample_events.append(
                    HMMEvent(
                        sample_id=sample_id,
                        map_name=map_name,
                        chromosome=chromosome,
                        event_index=0,
                        left_snp=left_marker.snp_name,
                        right_snp=right_marker.snp_name,
                        left_coordinate=left_marker.coordinate,
                        right_coordinate=right_marker.coordinate,
                        left_state=left_state,
                        right_state=right_state,
                        event_bp=(left_marker.coordinate + right_marker.coordinate) // 2,
                    )
                )

        sample_events.sort(key=lambda event: (event.chromosome, event.event_bp))
        grouped_by_chr: dict[int, list[HMMEvent]] = defaultdict(list)
        for event in sample_events:
            grouped_by_chr[event.chromosome].append(event)

        kept_events: list[HMMEvent] = []
        for chromosome in sorted(grouped_by_chr):
            kept_events.extend(collapse_close_double_cos(grouped_by_chr[chromosome]))

        kept_events.sort(key=lambda event: (event.chromosome, event.event_bp))
        for event_index, event in enumerate(kept_events, start=1):
            event.event_index = event_index
            events.append(event)

    return events


def write_hmm_events(events: list[HMMEvent]) -> None:
    HMM_EVENTS_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    with HMM_EVENTS_OUTPUT.open("w", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(
            [
                "sample_id",
                "map",
                "chromosome",
                "event_index",
                "event_bp",
                "left_coordinate",
                "right_coordinate",
                "left_snp",
                "right_snp",
                "left_state",
                "right_state",
            ]
        )
        for event in events:
            writer.writerow(
                [
                    event.sample_id,
                    event.map_name,
                    event.chromosome,
                    event.event_index,
                    event.event_bp,
                    event.left_coordinate,
                    event.right_coordinate,
                    event.left_snp,
                    event.right_snp,
                    event.left_state,
                    event.right_state,
                ]
            )


def write_qc_summary(rows: list[dict[str, int | str]]) -> None:
    QC_SUMMARY_OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    fieldnames = [
        "map",
        "input_markers",
        "deduplicated_markers",
        "removed_duplicate_markers",
        "input_samples",
        "kept_samples",
        "removed_samples_missingness",
        "removed_markers_missingness",
        "removed_markers_allele_balance",
        "removed_markers_isolated_flip",
        "final_markers",
        "hmm_events",
    ]
    with QC_SUMMARY_OUTPUT.open("w", newline="") as handle:
        writer = csv.DictWriter(handle, delimiter="\t", fieldnames=fieldnames)
        writer.writeheader()
        writer.writerows(rows)


def main() -> None:
    os.environ.setdefault("MPLCONFIGDIR", "/tmp/finemap-mpl")
    MATRIX_DIR.mkdir(parents=True, exist_ok=True)

    markers_by_map = load_markers()
    all_events: list[HMMEvent] = []
    qc_rows: list[dict[str, int | str]] = []

    for map_name, markers in sorted(markers_by_map.items()):
        clean_markers, _, summary = clean_population(markers)
        write_clean_matrix(map_name, clean_markers)
        hmm_events = build_hmm_events(map_name, clean_markers)
        all_events.extend(hmm_events)
        qc_rows.append(
            {
                "map": map_name,
                "input_markers": summary["input_markers"],
                "deduplicated_markers": summary["deduplicated_markers"],
                "removed_duplicate_markers": summary["removed_duplicate_markers"],
                "input_samples": summary["input_samples"],
                "kept_samples": summary["kept_samples"],
                "removed_samples_missingness": summary["removed_samples_missingness"],
                "removed_markers_missingness": summary["removed_markers_missingness"],
                "removed_markers_allele_balance": summary["removed_markers_allele_balance"],
                "removed_markers_isolated_flip": summary["removed_markers_isolated_flip"],
                "final_markers": summary["final_markers"],
                "hmm_events": len(hmm_events),
            }
        )

    write_hmm_events(all_events)
    write_qc_summary(qc_rows)
    print(
        f"Wrote {len(all_events)} HMM-based CO intervals to {HMM_EVENTS_OUTPUT}, "
        f"cleaned matrices to {MATRIX_DIR}, and QC summary to {QC_SUMMARY_OUTPUT}"
    )


if __name__ == "__main__":
    main()
