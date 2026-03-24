#!/usr/bin/env python3

from __future__ import annotations

import csv
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from openpyxl import load_workbook


INPUT_XLSX = Path("/Users/jeffreyross-ibarra/Downloads/gb-2013-14-9-r103-S4.xlsx")
LONG_OUTPUT = Path("results/co_events_long.tsv")
LONG_FILTERED_OUTPUT = Path("results/co_events_long_filtered.tsv")
WIDE_OUTPUT = Path("results/co_events_wide.tsv")

MIN_SUPPORT_MARKERS = 2
MIN_SUPPORT_BP = 1_000_000
DOUBLE_CO_MAX_BP = 2_000_000


@dataclass
class Marker:
    chromosome: int
    coordinate: int
    snp_name: str
    genotype: str


@dataclass
class Event:
    sample_id: str
    map_name: str
    chromosome: int
    event_index: int
    left_snp: str
    right_snp: str
    left_coordinate: int
    right_coordinate: int
    left_genotype: str
    right_genotype: str
    event_bp: int


@dataclass
class Run:
    genotype: str
    markers: list[Marker]

    @property
    def marker_count(self) -> int:
        return len(self.markers)

    @property
    def span_bp(self) -> int:
        return self.markers[-1].coordinate - self.markers[0].coordinate

    @property
    def start_marker(self) -> Marker:
        return self.markers[0]

    @property
    def end_marker(self) -> Marker:
        return self.markers[-1]


def load_markers() -> dict[str, list[Marker]]:
    wb = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    ws = wb["Table_S3"]
    header = list(next(ws.iter_rows(min_row=3, max_row=3, values_only=True)))
    idx = {name: i for i, name in enumerate(header)}

    markers_by_map: dict[str, list[Marker]] = defaultdict(list)

    for row in ws.iter_rows(min_row=4, values_only=True):
        map_name = row[idx["map"]]
        snp_name = row[idx["SNP_name"]]
        chromosome = row[idx["chr_phy"]]
        coordinate = row[idx["coordinate"]]
        raw_data = row[idx["raw_data"]]

        if not map_name or not snp_name or raw_data is None:
            continue
        if chromosome in (None, 0) or coordinate is None:
            continue
        if len(raw_data) <= 6:
            continue

        markers_by_map[map_name].append(
            Marker(
                chromosome=int(chromosome),
                coordinate=int(coordinate),
                snp_name=str(snp_name),
                genotype=str(raw_data),
            )
        )

    return markers_by_map


def build_runs(markers: list[Marker], sample_idx: int) -> list[Run]:
    runs: list[Run] = []
    current: Run | None = None

    for marker in markers:
        if sample_idx >= len(marker.genotype):
            continue

        genotype = marker.genotype[sample_idx]
        if genotype not in {"A", "B"}:
            continue

        if current is None or current.genotype != genotype:
            current = Run(genotype=genotype, markers=[marker])
            runs.append(current)
            continue

        current.markers.append(marker)

    return runs


def event_from_transition(sample_id: str, map_name: str, event_index: int, previous_run: Run, next_run: Run) -> Event:
    left_marker = previous_run.end_marker
    right_marker = next_run.start_marker
    return Event(
        sample_id=sample_id,
        map_name=map_name,
        chromosome=left_marker.chromosome,
        event_index=event_index,
        left_snp=left_marker.snp_name,
        right_snp=right_marker.snp_name,
        left_coordinate=left_marker.coordinate,
        right_coordinate=right_marker.coordinate,
        left_genotype=previous_run.genotype,
        right_genotype=next_run.genotype,
        event_bp=(left_marker.coordinate + right_marker.coordinate) // 2,
    )


def filter_close_double_cos(events: list[Event]) -> list[Event]:
    keep = [True] * len(events)
    idx = 0
    while idx < len(events) - 1:
        current = events[idx]
        nxt = events[idx + 1]
        if nxt.event_bp - current.event_bp <= DOUBLE_CO_MAX_BP:
            keep[idx] = False
            keep[idx + 1] = False
            idx += 2
            continue
        idx += 1
    return [event for event, keep_flag in zip(events, keep) if keep_flag]


def build_raw_events(markers_by_map: dict[str, list[Marker]]) -> list[Event]:
    events: list[Event] = []

    for map_name, markers in sorted(markers_by_map.items()):
        markers.sort(key=lambda m: (m.chromosome, m.coordinate, m.snp_name))
        sample_count = max(len(marker.genotype) for marker in markers)
        sample_event_counts = defaultdict(int)

        for sample_idx in range(sample_count):
            sample_id = f"{map_name}_ind{sample_idx + 1:03d}"
            previous_by_chr: dict[int, tuple[str, Marker]] = {}

            for marker in markers:
                if sample_idx >= len(marker.genotype):
                    continue

                genotype = marker.genotype[sample_idx]
                if genotype not in {"A", "B"}:
                    continue

                previous = previous_by_chr.get(marker.chromosome)
                if previous is None:
                    previous_by_chr[marker.chromosome] = (genotype, marker)
                    continue

                previous_genotype, previous_marker = previous
                if genotype != previous_genotype:
                    sample_event_counts[sample_id] += 1
                    events.append(
                        Event(
                            sample_id=sample_id,
                            map_name=map_name,
                            chromosome=marker.chromosome,
                            event_index=sample_event_counts[sample_id],
                            left_snp=previous_marker.snp_name,
                            right_snp=marker.snp_name,
                            left_coordinate=previous_marker.coordinate,
                            right_coordinate=marker.coordinate,
                            left_genotype=previous_genotype,
                            right_genotype=genotype,
                            event_bp=(previous_marker.coordinate + marker.coordinate) // 2,
                        )
                    )

                previous_by_chr[marker.chromosome] = (genotype, marker)

    return events


def build_filtered_events(markers_by_map: dict[str, list[Marker]]) -> list[Event]:
    filtered_events: list[Event] = []

    for map_name, markers in sorted(markers_by_map.items()):
        markers.sort(key=lambda m: (m.chromosome, m.coordinate, m.snp_name))
        sample_count = max(len(marker.genotype) for marker in markers)

        markers_by_chromosome: dict[int, list[Marker]] = defaultdict(list)
        for marker in markers:
            markers_by_chromosome[marker.chromosome].append(marker)

        for sample_idx in range(sample_count):
            sample_id = f"{map_name}_ind{sample_idx + 1:03d}"
            sample_events: list[Event] = []

            for chromosome, chromosome_markers in sorted(markers_by_chromosome.items()):
                runs = build_runs(chromosome_markers, sample_idx)
                if len(runs) < 2:
                    continue

                for previous_run, next_run in zip(runs, runs[1:]):
                    if next_run.marker_count < MIN_SUPPORT_MARKERS:
                        continue
                    if next_run.span_bp < MIN_SUPPORT_BP:
                        continue
                    sample_events.append(
                        event_from_transition(
                            sample_id=sample_id,
                            map_name=map_name,
                            event_index=0,
                            previous_run=previous_run,
                            next_run=next_run,
                        )
                    )

            sample_events.sort(key=lambda event: (event.chromosome, event.event_bp))

            grouped_by_chr: dict[int, list[Event]] = defaultdict(list)
            for event in sample_events:
                grouped_by_chr[event.chromosome].append(event)

            kept_events: list[Event] = []
            for chromosome in sorted(grouped_by_chr):
                kept_events.extend(filter_close_double_cos(grouped_by_chr[chromosome]))

            kept_events.sort(key=lambda event: (event.chromosome, event.event_bp))
            for event_index, event in enumerate(kept_events, start=1):
                event.event_index = event_index
                filtered_events.append(event)

    return filtered_events


def write_long(events: list[Event], output_path: Path) -> None:
    output_path.parent.mkdir(parents=True, exist_ok=True)
    with output_path.open("w", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(
            [
                "sample_id",
                "map",
                "chromosome",
                "event_index",
                "event_bp",
                "left_coordinate",
                "right_coordinate",
                "left_snp",
                "right_snp",
                "left_genotype",
                "right_genotype",
            ]
        )
        for event in events:
            writer.writerow(
                [
                    event.sample_id,
                    event.map_name,
                    event.chromosome,
                    event.event_index,
                    event.event_bp,
                    event.left_coordinate,
                    event.right_coordinate,
                    event.left_snp,
                    event.right_snp,
                    event.left_genotype,
                    event.right_genotype,
                ]
            )


def write_wide(events: list[Event]) -> None:
    events_by_sample: dict[str, list[str]] = defaultdict(list)
    for event in events:
        events_by_sample[event.sample_id].append(f"chr{event.chromosome}:{event.event_bp}")

    sample_ids = sorted(events_by_sample)
    max_events = max((len(values) for values in events_by_sample.values()), default=0)

    with WIDE_OUTPUT.open("w", newline="") as handle:
        writer = csv.writer(handle, delimiter="\t")
        writer.writerow(sample_ids)
        for row_idx in range(max_events):
            row = []
            for sample_id in sample_ids:
                values = events_by_sample[sample_id]
                row.append(values[row_idx] if row_idx < len(values) else "")
            writer.writerow(row)


def main() -> None:
    markers_by_map = load_markers()
    raw_events = build_raw_events(markers_by_map)
    filtered_events = build_filtered_events(markers_by_map)
    write_long(raw_events, LONG_OUTPUT)
    write_long(filtered_events, LONG_FILTERED_OUTPUT)
    write_wide(raw_events)
    print(
        f"Wrote {len(raw_events)} raw events to {LONG_OUTPUT}, "
        f"{len(filtered_events)} filtered events to {LONG_FILTERED_OUTPUT}, "
        f"and the raw wide table to {WIDE_OUTPUT}"
    )


if __name__ == "__main__":
    main()
